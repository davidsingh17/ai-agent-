from fastapi import APIRouter, UploadFile, File, HTTPException, Query
import uuid
import os
import mimetypes
from typing import List, Optional
from uuid import UUID
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.services.storage import upload_bytes
from app.services.repository_invoices import insert_invoice_header, insert_invoice_lines
from app.services.parsers.invoice_xml import parse_xml_fatturapa
from app.services.parsers.invoice_pdf import parse_pdf_invoice

from app.schemas.invoice import (
    InvoiceOut, InvoiceListItem, InvoiceListResponse, PresignedUrlOut
)
from app.services.invoice_service import (
    list_invoices, get_invoice, get_presigned_url
)

from fastapi import Response
from fastapi.responses import StreamingResponse
from io import BytesIO, StringIO  # 👈 aggiunto StringIO
import csv
from xml.etree.ElementTree import Element, SubElement, tostring
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

router = APIRouter(prefix="/invoices", tags=["invoices"])
IS_TESTING = os.getenv("TESTING") == "1"


def _merge_defaults(parsed: dict) -> dict:
    fields = parsed.get("fields", {}) or {}
    fields.setdefault("valuta", "EUR")
    parsed["fields"] = fields
    parsed.setdefault("righe", [])
    return parsed


def _to_float_safe(x) -> Optional[float]:
    try:
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).replace(".", "").replace(",", ".").replace("%", "").strip()
        return float(s)
    except Exception:
        return None


def _backfill_amounts(fields: dict) -> dict:
    imp = _to_float_safe(fields.get("imponibile"))
    iva = _to_float_safe(fields.get("iva"))
    tot = _to_float_safe(fields.get("totale"))

    if imp is None and iva is not None and tot is not None:
        base = round(tot - iva, 2)
        if base >= 0:
            imp = base
    elif iva is None and imp is not None and tot is not None:
        tax = round(tot - imp, 2)
        if tax >= 0:
            iva = tax
    elif tot is None and imp is not None and iva is not None:
        tot = round(imp + iva, 2)

    if imp is not None and iva is not None:
        calc = round(imp + iva, 2)
        if tot is None or abs(calc - tot) <= 0.05:
            tot = calc

    fields["imponibile"] = imp
    fields["iva"] = iva
    fields["totale"] = tot
    return fields


@router.post("/extract", response_model=InvoiceOut)
async def extract_invoice(file: UploadFile = File(...)):
    try:
        file_bytes = await file.read()

        name_lower = (file.filename or "").lower()
        content_type = (file.content_type or "").lower()
        if name_lower.endswith(".xml") or "xml" in content_type:
            parsed = parse_xml_fatturapa(file_bytes)
        elif name_lower.endswith(".pdf") or "pdf" in content_type:
            parsed = parse_pdf_invoice(file_bytes)
        else:
            parsed = {"fields": {"valuta": "EUR"}, "righe": []}

        parsed = _merge_defaults(parsed)
        parsed["fields"] = _backfill_amounts(parsed["fields"])
        f = parsed["fields"]

        fake_s3 = {"bucket": "test-bucket", "key": f"invoices/{uuid.uuid4()}_{file.filename or 'file'}"}
        upload_result = None
        try:
            file_id = str(uuid.uuid4())
            s3_key = f"invoices/{file_id}_{file.filename}"
            upload_result = upload_bytes(s3_key, file_bytes, content_type=(file.content_type or "application/octet-stream"))
        except Exception:
            if not IS_TESTING:
                raise
            upload_result = fake_s3

        invoice_uuid = str(uuid.uuid4())
        try:
            insert_invoice_header(
                id=invoice_uuid,
                s3_bucket=upload_result["bucket"],
                s3_key=upload_result["key"],
                filename=file.filename,
                invoice_number=f.get("invoice_number"),
                intestatario=f.get("intestatario"),
                partita_iva=f.get("partita_iva"),
                codice_fiscale=f.get("codice_fiscale"),
                issue_date=f.get("data_emissione"),
                due_date=f.get("data_scadenza"),
                currency=f.get("valuta", "EUR"),
                imponibile=f.get("imponibile"),
                iva=f.get("iva"),
                totale=f.get("totale"),
            )
            insert_invoice_lines(invoice_id=invoice_uuid, lines=parsed.get("righe", []))
        except Exception:
            if not IS_TESTING:
                raise

        return InvoiceOut(
            id=invoice_uuid,
            s3=upload_result,
            filename=file.filename,
            fields=f,
            righe=parsed.get("righe", []),
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("", response_model=InvoiceListResponse)
def list_invoices_route(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    q: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
    order_by: Optional[str] = Query("created_at"),
    order_dir: Optional[str] = Query("desc"),
):
    items_raw, total = list_invoices(
        limit=limit, offset=offset, q=q, date_from=date_from, date_to=date_to, order_by=order_by or "created_at", order_dir=order_dir or "desc"
    )
    items: List[InvoiceListItem] = []
    for inv in items_raw:
        fields = inv.get("fields", {}) or {}
        items.append(InvoiceListItem(
            id=inv["id"],
            filename=inv.get("filename", "document.pdf"),
            intestatario=fields.get("intestatario"),
            invoice_number=fields.get("invoice_number"),
            data_emissione=fields.get("data_emissione"),
            totale=fields.get("totale"),
        ))
    return InvoiceListResponse(items=items, total=total, limit=limit, offset=offset)


@router.get("/{invoice_id}", response_model=InvoiceOut)
def get_invoice_route(invoice_id: UUID):
    inv = get_invoice(str(invoice_id))
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return InvoiceOut(
        id=inv["id"],
        s3=inv["s3"],
        filename=inv.get("filename", "document.pdf"),
        fields=inv.get("fields", {}) or {},
        righe=inv.get("righe", []) or [],
    )


@router.get("/{invoice_id}/download", response_model=PresignedUrlOut)
def download_invoice_route(
    invoice_id: UUID,
    expires_in: int = Query(900, ge=60, le=86400)
):
    inv = get_invoice(str(invoice_id))
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    url = get_presigned_url(bucket=inv["s3"]["bucket"], key=inv["s3"]["key"], expires_in=expires_in)
    if not url:
        raise HTTPException(status_code=500, detail="Unable to generate presigned URL")
    return PresignedUrlOut(url=url, expires_in=expires_in)


@router.get("/{invoice_id}/preview", response_model=PresignedUrlOut)
def preview_invoice_route(
    invoice_id: UUID,
    expires_in: int = Query(900, ge=60, le=86400)
):
    inv = get_invoice(str(invoice_id))
    if not inv:
        raise HTTPException(status_code=404, detail="Unable to find invoice")

    filename = inv.get("filename") or ""
    key = inv["s3"]["key"]
    guessed, _ = mimetypes.guess_type(filename or key)
    content_type = guessed or "application/octet-stream"

    url = get_presigned_url(
        bucket=inv["s3"]["bucket"],
        key=key,
        expires_in=expires_in,
        inline=True,
        content_type=content_type,
        filename=filename,
    )
    if not url:
        raise HTTPException(status_code=500, detail="Unable to generate preview URL")
    return PresignedUrlOut(url=url, expires_in=expires_in)


# ========= ESPORTAZIONI =========

def _normalize_field(v):
    return "" if v is None else str(v)


@router.get("/export.csv")
@router.get("/export/csv")  # 👈 alias senza punto (compatibile con UI)
def export_invoices_csv(
    limit: int = Query(1000, ge=1, le=5000),
    offset: int = Query(0, ge=0),
    q: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
    order_by: Optional[str] = Query("created_at"),
    order_dir: Optional[str] = Query("desc"),
    # 👇 opzioni per Excel
    sep: str = Query(";", min_length=1, max_length=1, description="Separatore CSV"),
    bom: int = Query(1, ge=0, le=1, description="Scrivi BOM UTF-8 (1=on)"),
):
    items_raw, _ = list_invoices(
        limit=limit, offset=offset, q=q,
        date_from=date_from, date_to=date_to,
        order_by=order_by or "created_at",
        order_dir=order_dir or "desc",
    )

    def generate():
        # BOM per Excel
        if bom:
            yield b"\xef\xbb\xbf"

        sio = StringIO(newline="")
        writer = csv.writer(sio, delimiter=sep, lineterminator="\r\n")

        # header
        writer.writerow(["id", "filename", "intestatario", "invoice_number", "data_emissione", "totale"])
        yield sio.getvalue().encode("utf-8")
        sio.seek(0); sio.truncate(0)

        # righe
        for inv in items_raw:
            f = inv.get("fields", {}) or {}
            writer.writerow([
                _normalize_field(inv.get("id")),
                _normalize_field(inv.get("filename")),
                _normalize_field(f.get("intestatario")),
                _normalize_field(f.get("invoice_number")),
                _normalize_field(f.get("data_emissione")),
                _normalize_field(f.get("totale")),
            ])
            yield sio.getvalue().encode("utf-8")
            sio.seek(0); sio.truncate(0)

    headers = {"Content-Disposition": 'attachment; filename="invoices.csv"'}
    return StreamingResponse(generate(), media_type="text/csv; charset=utf-8", headers=headers)


@router.get("/export.xml")
@router.get("/export/xml")  # 👈 alias senza punto
def export_invoices_xml(
    limit: int = Query(1000, ge=1, le=5000),
    offset: int = Query(0, ge=0),
    q: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    order_by: Optional[str] = Query("created_at"),
    order_dir: Optional[str] = Query("desc"),
):
    items_raw, _ = list_invoices(limit=limit, offset=offset, q=q, date_from=date_from, date_to=date_to, order_by=order_by or "created_at", order_dir=order_dir or "desc")

    root = Element("Invoices")
    for inv in items_raw:
        fields = inv.get("fields", {}) or {}
        xinv = SubElement(root, "Invoice", {"id": str(inv.get("id") or "")})
        SubElement(xinv, "Filename").text = _normalize_field(inv.get("filename"))
        xf = SubElement(xinv, "Fields")
        for k in ["intestatario", "invoice_number", "data_emissione", "data_scadenza",
                  "partita_iva", "codice_fiscale", "valuta", "imponibile", "iva", "totale"]:
            SubElement(xf, k).text = _normalize_field(fields.get(k))

    xml_bytes = tostring(root, encoding="utf-8", method="xml")
    headers = {"Content-Disposition": 'attachment; filename="invoices.xml"'}
    return Response(content=xml_bytes, media_type="application/xml", headers=headers)


@router.get("/{invoice_id}/export.pdf")
def export_invoice_pdf(invoice_id: UUID):
    inv = get_invoice(str(invoice_id))
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")

    fields = inv.get("fields", {}) or {}

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    y = h - 50
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "Fattura — Riepilogo")
    y -= 25

    c.setFont("Helvetica", 11)

    def line(label, value):
        nonlocal y
        c.drawString(40, y, f"{label}: {value or '-'}")
        y -= 18

    line("ID", inv["id"])
    line("Filename", inv.get("filename"))
    line("Intestatario", fields.get("intestatario"))
    line("Partita IVA", fields.get("partita_iva"))
    line("Codice Fiscale", fields.get("codice_fiscale"))
    line("Numero Fattura", fields.get("invoice_number"))
    line("Data Emissione", fields.get("data_emissione"))
    line("Data Scadenza", fields.get("data_scadenza"))
    line("Valuta", fields.get("valuta") or "EUR")
    line("Imponibile", fields.get("imponibile"))
    line("IVA", fields.get("iva"))
    line("Totale", fields.get("totale"))

    c.showPage()
    c.save()
    buf.seek(0)

    headers = {"Content-Disposition": f'attachment; filename="invoice_{inv["id"]}.pdf"'}
    return StreamingResponse(buf, media_type="application/pdf", headers=headers)


@router.get("/export/xlsx")
def export_invoices_xlsx(
    limit: int = Query(1000, ge=1, le=5000),
    offset: int = Query(0, ge=0),
    q: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
    order_by: str = Query("created_at"),
    order_dir: str = Query("desc"),
):
    items_raw, _total = list_invoices(
        limit=limit, offset=offset, q=q,
        date_from=date_from, date_to=date_to,
        order_by=order_by or "created_at",
        order_dir=order_dir or "desc",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Fatture"

    headers = ["id", "filename", "intestatario", "invoice_number", "data_emissione", "totale"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    def _parse_date(s: Optional[str]):
        if not s:
            return None
        try:
            return datetime.fromisoformat(s).date()
        except Exception:
            return s

    for inv in items_raw:
        f = inv.get("fields", {}) or {}
        ws.append([
            inv.get("id"),
            inv.get("filename"),
            f.get("intestatario"),
            f.get("invoice_number"),
            _parse_date(f.get("data_emissione")),
            f.get("totale"),
        ])

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=6).number_format = "#,##0.00"

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[letter].width = min(max(12, maxlen + 2), 48)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'invoices_{datetime.utcnow().strftime("%Y-%m-%d_%H-%M")}.xlsx'
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
